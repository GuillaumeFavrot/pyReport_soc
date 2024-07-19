from numpy import ndarray
import re
from openpyxl import load_workbook
from utilities.data_export.tools.column_to_number import column_to_number
from utilities.timer.timer import Timer
from utilities.logger.logger import Logger

def export_data(timer:Timer, logger:Logger, df: ndarray, output_file_name: str, config:list[tuple]) -> None:
    """Exporte l'une des colonnes spécifée d'un dataframe dans l'une feuille d'un fichier excel donné à une position précise (au format A1,B2, etc...)"""

    logger.info(f"{timer.get_time()} - Export des données d'analyse vers {output_file_name}")
    
    # Vérification du type de la variable "output_file_name"
    if type(output_file_name) != str:
        logger.error(f'{timer.get_time()} - Le nom de la variable "output_file_name" doit être chaînes de caractères')
        raise TypeError("File name must be a string")
    
    # Vérification du type de la variable "config"
    if type(config) != list:
        logger.error(f'{timer.get_time()} - La variable "config" doit être une liste')
        raise TypeError("config varaible must be a list")
    
    # Vérication du type des éléments de la variable "config" et de leur structure
    for e in config:
        if type(e) != tuple:
            logger.error(f'{timer.get_time()} - Les éléments de la variable "config" doivent être des tuples')
            raise TypeError("Elements of the config variable must be tuples")
        
        if len(e) != 3:
            logger.error(f'{timer.get_time()} - Les éléments de la variable "config" doivent contenir trois valeurs')
            raise Exception("Elements of the config variable must contain three values")
        
        for v in e:
            if type(v) != str:
                logger.error(f'{timer.get_time()} - Les éléments de la variable "config" doivent être des chaînes de caractères')
                raise TypeError("Elements of the config variable must be strings") 

    #Création de l'adresse du fichier
    file_path = f"./output/{output_file_name}"

    #Ouverture du fichier destination
    logger.info(f"{timer.get_time()} - Ouverture du fichier")
    wb = load_workbook(file_path)
    logger.info(f"{timer.get_time()} - Fichier chargé")

    for c in config:
        #Ectraction des donées de configuration
        (column_name, sheet_name, cell) = c
        
        logger.info(f"{timer.get_time()} - Copie des données {column_name} dans la feuille {sheet_name} du fichier {output_file_name}")
        
        #Création des variables necessaires à l'exportation
        row = int(re.sub("[^ 0-9]+", '', cell))
        column = column_to_number(re.sub("[^ a-zA-Z]+", '', cell))
        data = df.get(column_name).tolist()

        #Sélection de la feuille
        ws = wb[sheet_name]

        #Copiage des données
        for i, value in enumerate(data, start=row):
            cell = ws.cell(row=i, column=column).value=value

    #Sauvegarde du fichier
    logger.info(f"{timer.get_time()} - Sauvegarde en cours")
    wb.save(file_path)
    logger.info(f"{timer.get_time()} - Sauvegarde terminée")
   

